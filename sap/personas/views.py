from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook
from rest_framework import viewsets, permissions

from personas.forms import PersonaFormulario
from personas.models import Persona, Curso
from personas.serializers import PersonaSerializer, CursoSerializer


# Create your views here.

# PersonaFormulario = modelform_factory(Persona, exclude=['activo',])
def agregar_persona(request):
    pagina = loader.get_template('personas/agregar.html')
    if request.method == 'GET':
        formulario = PersonaFormulario
    elif request.method == 'POST':
        formulario = PersonaFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario':formulario}
    return HttpResponse(pagina.render(datos,request))

def modificar_persona(request, id):
    pagina = loader.get_template('personas/modificar.html')
    persona = get_object_or_404(Persona, pk=id)
    if request.method == 'GET':
        formulario = PersonaFormulario(instance=persona)
    elif request.method == 'POST':
        formulario = PersonaFormulario(request.POST, instance=persona)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_persona(request, id):
    # persona = Persona.objects.get(pk=id)
    persona = get_object_or_404(Persona, pk=id)
    datos = {'persona':persona}
    # print(persona)
    pagina = loader.get_template('personas/ver.html')
    return HttpResponse(pagina.render(datos, request))

def eliminar_persona(request, id):
    persona = get_object_or_404(Persona, pk=id)
    if persona:
        persona.delete()
        return redirect('inicio')


# Nuestra clase hereda de la vista genérica TemplateView

def generar_reporte(request, *args, **kwargs):
    # Obtenemos todas las personas de nuestra base de datos
    personas = Persona.objects.order_by('apellido', 'nombre')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE PERSONAS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'ID'
    ws['C3'] = 'NOMBRE'
    ws['D3'] = 'APELLIDO'
    ws['E3'] = 'EMAIL'
    ws['F3'] = 'CURSO'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for persona in personas:
        ws.cell(row=cont, column=2).value = persona.id
        ws.cell(row=cont, column=3).value = persona.nombre
        ws.cell(row=cont, column=4).value = persona.apellido
        ws.cell(row=cont, column=5).value = persona.email
        ws.cell(row=cont, column=6).value = persona.curso.nombre
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReportePersonasExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


class PersonaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Persona.objects.all().order_by('-apellido')
    serializer_class = PersonaSerializer
    permission_classes = [permissions.IsAuthenticated]


class CursoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Curso.objects.all()
    serializer_class = CursoSerializer
    permission_classes = [permissions.IsAuthenticated]